from datetime import timedelta
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.utils import timezone
from io import BytesIO
from unittest.mock import patch
import openpyxl

from core.models import Modalidade, Pessoa, Processo, ProcessoItem, Secretaria, StatusProcesso
from docs.models import ProcessoAnexo
from core.utils.bll_export import export_bll_xlsx
from core.utils.bll_import import import_bll_file
from workflow.models import DocumentoProcessoWorkflow, PNCPDetalhamentoFila, ProcessoWorkflow, ModuloSistema
from workflow.views import _build_texto_pdf, _normalizar_pdf_para_padrao_documental
from workflow.services.pncp_queue import enqueue_pncp_detalhamento, processar_fila_pncp
from workflow.services.pncp_publish import PNCPPublishClient


class SmokeAuthTest(TestCase):
    def test_logout_encerra_sessao(self):
        user = get_user_model().objects.create_user(username='smoke', password='senha-smoke')
        self.assertTrue(self.client.login(username='smoke', password='senha-smoke'))

        response = self.client.post('/sirel/logout/')
        self.assertIn(response.status_code, (200, 302))

        home = self.client.get('/sirel/')
        self.assertEqual(home.status_code, 302)
        self.assertIn('/sirel/login/', home['Location'])


class SmokeBLLImportExportTest(TestCase):
    def setUp(self):
        self.processo = Processo.objects.create(
            numero_processo_adm='SMK-001',
            numero_edital='PE-001',
            ano_referencia=2026,
            objeto='Teste smoke BLL',
        )

    def test_importa_multisection_e_exporta_layout(self):
        content = """<PARTICIPANTES>
2|11111111000191|FORNECEDOR A|RUA A|CENTRO||45999000|CIDADE A|BA
<PROCESSO>
PE-001-2026
<LOTES>
1|HOMOLOGADO|GLOBAL|1
<VALORES UNITARIOS>
1|1|ITEM TESTE|10,0000|UND|5,0000|MARCA X|MODELO Y|5,0000|4,0000
<CLASSIFICACAO>
1|1|FORNECEDOR A|11111111000191|40,0000|1|1|1
"""
        up = SimpleUploadedFile('bll.csv', content.encode('utf-8'), content_type='text/csv')
        res = import_bll_file(self.processo, up)
        self.assertEqual(res['mode'], 'bll_pipe_multisection')

        self.assertEqual(self.processo.itens.count(), 1)
        item = self.processo.itens.first()
        self.assertEqual(item.numero_item, 1)
        self.assertEqual(item.lote.numero, 1)
        self.assertEqual(item.ofertas.count(), 1)

        buf = BytesIO()
        export_bll_xlsx(self.processo, buf)
        buf.seek(0)
        wb = openpyxl.load_workbook(buf, data_only=True)
        self.assertEqual(wb.sheetnames, ['LOTES', 'ITENS', 'TIPOLANCE'])
        ws_lotes = wb['LOTES']
        ws_itens = wb['ITENS']
        self.assertEqual([ws_lotes.cell(1, c).value for c in range(1, 9)], ['LOTE', 'TÍTULO', 'TIPO LANCE', 'QUANTIDADE', 'MARGEM LANCE', 'GARANTIA', 'LOCAL ENTREGA', 'EXCLUSIVO ME'])
        self.assertEqual([ws_itens.cell(1, c).value for c in range(1, 9)], ['LOTE', 'ITEM', 'DESCRIÇÃO', 'UNID', 'QUANTIDADE', 'VALOR DE REFERÊNCIA', 'INFO DETALHADA', 'ARQUIVO REQUERIDO'])


class SmokePNCPPublishOptionalTest(TestCase):
    def setUp(self):
        self.processo = Processo.objects.create(
            numero_processo_adm='SMK-PNCP',
            numero_edital='PE-009',
            ano_referencia=2026,
            objeto='Teste envio opcional PNCP',
        )
        ProcessoItem.objects.create(
            processo=self.processo,
            numero_item=1,
            descricao_snapshot='Item smoke',
            unidade_snapshot='UND',
            quantidade=1,
        )

    def test_envio_desabilitado_retorna_skip(self):
        client = PNCPPublishClient()
        res = client.enviar(self.processo)
        self.assertEqual(res['status'], 'SKIPPED_DISABLED')

    @override_settings(PNCP_ENVIO_HABILITADO=True, PNCP_ENVIO_DRY_RUN=True)
    def test_envio_habilitado_em_simulacao(self):
        client = PNCPPublishClient()
        res = client.enviar(self.processo)
        self.assertEqual(res['status'], 'DRY_RUN')


class SmokePNCPQueueTest(TestCase):
    def setUp(self):
        self.processo = Processo.objects.create(
            numero_processo_adm='SMK-FILA',
            ano_referencia=2026,
            objeto='Teste fila detalhamento PNCP',
        )

    def test_enqueue_reseta_status_para_pendente(self):
        fila = PNCPDetalhamentoFila.objects.create(
            processo=self.processo,
            numero_controle_pncp='12345678901234-1-1/2026',
            status=PNCPDetalhamentoFila.Status.ERRO,
            ultimo_erro='falha anterior',
            prioridade=90,
            origem='IMPORTACAO_RAPIDA',
        )
        fila_atu, criado = enqueue_pncp_detalhamento(
            self.processo,
            numero_controle='12345678901234-1-1/2026',
            origem='IMPORTACAO_RAPIDA',
            prioridade=20,
        )
        self.assertFalse(criado)
        self.assertEqual(fila_atu.id, fila.id)
        fila_atu.refresh_from_db()
        self.assertEqual(fila_atu.status, PNCPDetalhamentoFila.Status.PENDENTE)
        self.assertEqual(fila_atu.prioridade, 20)

    def test_processar_fila_conclui_job(self):
        enqueue_pncp_detalhamento(
            self.processo,
            numero_controle='12345678901234-1-1/2026',
            origem='IMPORTACAO_RAPIDA',
            prioridade=10,
        )
        with patch(
            'workflow.views._reprocessar_pncp_processo',
            return_value={
                'numero_controle': '12345678901234-1-1/2026',
                'sync': {'itens_importados': 1, 'fornecedores_atualizados': 1},
                'itens_detalhados': 1,
                'resultados_detalhados': 1,
                'erros': [],
            },
        ):
            resumo = processar_fila_pncp(limit=5)
        self.assertEqual(resumo['capturados'], 1)
        self.assertEqual(resumo['concluidos'], 1)
        self.assertEqual(resumo['parciais'], 0)
        self.assertEqual(resumo['erros'], 0)

        fila = PNCPDetalhamentoFila.objects.get(processo=self.processo)
        self.assertEqual(fila.status, PNCPDetalhamentoFila.Status.CONCLUIDO)


class SmokePlanejamentoNumeracaoSirelTest(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username='planejamento-smoke',
            password='senha-smoke',
        )
        self.client.login(username='planejamento-smoke', password='senha-smoke')

    def test_planejamento_cria_numero_sirel_sequencial_e_ref_externa(self):
        response_1 = self.client.post(
            '/sirel/planejamento/novo/',
            {
                'ano_referencia': 2026,
                'numero_processo_externo': 'EXT-2026-001',
                'objeto': 'Aquisicao de teste 1',
            },
        )
        self.assertEqual(response_1.status_code, 302)
        proc_1 = Processo.objects.get(objeto='Aquisicao de teste 1')
        self.assertEqual(proc_1.numero_processo_sirel, '0001/2026')
        self.assertEqual(proc_1.numero_processo_adm, 'EXT-2026-001')

        response_2 = self.client.post(
            '/sirel/planejamento/novo/',
            {
                'ano_referencia': 2026,
                'numero_processo_externo': '',
                'objeto': 'Aquisicao de teste 2',
            },
        )
        self.assertEqual(response_2.status_code, 302)
        proc_2 = Processo.objects.get(objeto='Aquisicao de teste 2')
        self.assertEqual(proc_2.numero_processo_sirel, '0002/2026')
        self.assertEqual(proc_2.numero_processo_adm, '')

    def test_busca_planejamento_aceita_numero_sirel_e_externo(self):
        processo = Processo.objects.create(
            numero_processo_sirel='0099/2026',
            numero_processo_adm='PROC-EXTERNO-0099',
            ano_referencia=2026,
            objeto='Objeto teste busca',
        )
        ProcessoWorkflow.objects.create(
            processo=processo,
            modulo_atual=ModuloSistema.PLANEJAMENTO,
            etapa_atual='DFD',
            situacao='EM_ANDAMENTO',
        )

        resp_sirel = self.client.get('/sirel/planejamento/', {'campo': 'numero', 'q': '0099/2026'})
        self.assertContains(resp_sirel, 'Objeto teste busca')

        resp_externo = self.client.get('/sirel/planejamento/', {'campo': 'numero', 'q': 'PROC-EXTERNO-0099'})
        self.assertContains(resp_externo, 'Objeto teste busca')


class SmokeLicitacaoExternaChecklistTest(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username='licitacao-smoke',
            password='senha-smoke',
        )
        self.client.login(username='licitacao-smoke', password='senha-smoke')
        self.modalidade = Modalidade.objects.create(nome='Dispensa de licitação - Lei 14.133/2021')

    def test_registro_externo_aceita_confirmacao_com_pendencias_e_upload(self):
        payload_base = {
            'ano_referencia': 2026,
            'numero_processo_externo': 'EXT-LIC-001',
            'numero_edital': 'DLS-001/2026',
            'objeto': 'Teste processo externo na licitação',
            'modalidade': self.modalidade.id,
            'status': '',
            'secretaria': '',
            'data_publicacao': '',
            'valor_estimado': '0',
        }

        # Sem confirmação: não registra porque há documentos pendentes.
        resposta_sem_confirmacao = self.client.post('/sirel/licitacao/novo-externo/', payload_base)
        self.assertEqual(resposta_sem_confirmacao.status_code, 200)
        self.assertFalse(Processo.objects.filter(numero_processo_adm='EXT-LIC-001').exists())

        # Com confirmação e ao menos um anexo inicial: registra no módulo Licitação.
        payload_confirmado = dict(payload_base)
        payload_confirmado['confirmar_pendencias'] = 'on'
        payload_confirmado['arquivo_dfd'] = SimpleUploadedFile('dfd.pdf', b'dfd-smoke', content_type='application/pdf')
        resposta_confirmada = self.client.post('/sirel/licitacao/novo-externo/', payload_confirmado)
        self.assertEqual(resposta_confirmada.status_code, 302)

        processo = Processo.objects.get(numero_processo_adm='EXT-LIC-001')
        self.assertEqual(processo.workflow.modulo_atual, ModuloSistema.LICITACAO)
        self.assertEqual(processo.workflow.etapa_atual, 'LICITACAO - FASE INTERNA')
        self.assertTrue(
            DocumentoProcessoWorkflow.objects.filter(
                processo=processo,
                tipo_documento='LICITACAO_DOC::dfd',
            ).exists()
        )

        upload_parecer = self.client.post(
            f'/sirel/licitacao/{processo.id}/documentos/upload/',
            {
                'doc_codigo': 'parecer_juridico',
                'arquivo': SimpleUploadedFile('parecer.pdf', b'parecer-smoke', content_type='application/pdf'),
            },
        )
        self.assertEqual(upload_parecer.status_code, 302)
        self.assertTrue(
            DocumentoProcessoWorkflow.objects.filter(
                processo=processo,
                tipo_documento='LICITACAO_DOC::parecer_juridico',
            ).exists()
        )

    def test_ci_html_e_relatorio_pendencias_disponiveis(self):
        processo = Processo.objects.create(
            numero_processo_sirel='0001/2026',
            numero_processo_adm='EXT-LIC-REL-001',
            ano_referencia=2026,
            objeto='Processo para validar CI HTML e relatorio',
            modalidade=self.modalidade,
        )
        ProcessoWorkflow.objects.create(
            processo=processo,
            modulo_atual=ModuloSistema.LICITACAO,
            etapa_atual='LICITACAO - FASE INTERNA',
            situacao='EM_ANDAMENTO',
        )
        DocumentoProcessoWorkflow.objects.create(
            processo=processo,
            modulo=ModuloSistema.LICITACAO,
            tipo_documento='LICITACAO_DOC::dfd',
            arquivo=SimpleUploadedFile('dfd.pdf', b'dfd', content_type='application/pdf'),
            ordem_cronologica=1,
        )

        resp_ci = self.client.get(f'/sirel/licitacao/{processo.id}/documentos/ci_orcamento_reserva/html/')
        self.assertEqual(resp_ci.status_code, 200)
        self.assertContains(resp_ci, 'Solicitação de reserva orçamentária')
        self.assertContains(resp_ci, '3 - Licitação')

        resp_relatorio = self.client.get(f'/sirel/licitacao/{processo.id}/relatorio-pendencias/')
        self.assertEqual(resp_relatorio.status_code, 200)
        self.assertContains(resp_relatorio, 'Relatório de pendências documentais')
        self.assertContains(resp_relatorio, 'FASE DE JULGAMENTO')
        self.assertContains(resp_relatorio, 'PLANEJADO')


class SmokeDashboardsGeraisTest(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username='dash-smoke',
            password='senha-smoke',
        )
        self.client.login(username='dash-smoke', password='senha-smoke')

    def test_dashboard_exibe_agenda_calendario_e_ranking_condutor(self):
        secretaria = Secretaria.objects.create(sigla='SMPS', nome='Secretaria Municipal de Promocao Social')
        modalidade = Modalidade.objects.create(nome='Pregao')
        status = StatusProcesso.objects.create(nome='Em andamento')
        condutor = Pessoa.objects.create(nome='Jonatas Sousa', cargo='Pregoeiro', secretaria=secretaria)
        processo = Processo.objects.create(
            numero_processo_sirel='0099/2026',
            numero_processo_adm='ADM-0099',
            ano_referencia=2026,
            objeto='Processo smoke para agenda operacional',
            secretaria=secretaria,
            modalidade=modalidade,
            status=status,
            condutor_processo=condutor,
            data_hora_abertura=timezone.now() + timedelta(hours=2),
            fim_recolhimento_propostas=timezone.now() + timedelta(days=1),
        )
        ProcessoWorkflow.objects.create(
            processo=processo,
            modulo_atual=ModuloSistema.LICITACAO,
            etapa_atual='LICITACAO - FASE EXTERNA',
            situacao='EM_ANDAMENTO',
        )

        resp = self.client.get('/sirel/dashboards/')
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'Operacao do dia')
        self.assertContains(resp, 'Calendario de processos')
        self.assertContains(resp, 'Ranking por pregoeiro / condutor')
        self.assertContains(resp, 'Jonatas Sousa')
        self.assertContains(resp, '0099/2026')


class SmokeModuloDocumentosTest(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username='docs-smoke',
            password='senha-smoke',
        )
        self.client.login(username='docs-smoke', password='senha-smoke')

    def test_modulo_documentos_lista_processo_fora_do_modulo(self):
        processo = Processo.objects.create(
            numero_processo_sirel='0049/2026',
            numero_processo_adm='ADM-0049',
            ano_referencia=2026,
            objeto='Processo de teste para módulo documentos',
        )
        ProcessoWorkflow.objects.create(
            processo=processo,
            modulo_atual=ModuloSistema.LICITACAO,
            etapa_atual='LICITACAO - FASE INTERNA',
            situacao='EM_ANDAMENTO',
        )
        DocumentoProcessoWorkflow.objects.create(
            processo=processo,
            modulo=ModuloSistema.LICITACAO,
            tipo_documento='LICITACAO_DOC::dfd',
            arquivo=SimpleUploadedFile('dfd.pdf', b'%PDF-1.4 dfd', content_type='application/pdf'),
            ordem_cronologica=1,
        )
        ProcessoAnexo.objects.create(
            processo=processo,
            tipo='OUTROS',
            descricao='Anexo docs',
            arquivo=SimpleUploadedFile('anexo.txt', b'teste', content_type='text/plain'),
        )

        resp = self.client.get('/sirel/modulos/DOCUMENTOS/', {'q': '0049/2026'})
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, '0049/2026')
        self.assertContains(resp, 'Gerar processo integral (PDF)')
        self.assertContains(resp, 'Gerar e-TCM')
        self.assertContains(resp, f'/sirel/documentos/processo/{processo.id}/integral/pdf/')
        self.assertContains(resp, f'/sirel/documentos/processo/{processo.id}/etcm/')

    @patch('workflow.views._normalizar_pdf_para_padrao_documental')
    @patch('workflow.views._gerar_processo_consolidado_pdf')
    def test_rota_geracao_processo_integral_sem_restricao_de_modulo(self, mock_gerar, mock_normalizar):
        processo = Processo.objects.create(
            numero_processo_sirel='0050/2026',
            numero_processo_adm='ADM-0050',
            ano_referencia=2026,
            objeto='Processo para geração integral',
        )
        ProcessoWorkflow.objects.create(
            processo=processo,
            modulo_atual=ModuloSistema.COMPRAS,
            etapa_atual='COMPRAS - PESQUISA',
            situacao='EM_ANDAMENTO',
        )
        mock_gerar.return_value = (b'%PDF-1.4\n%mock\n', [])
        mock_normalizar.return_value = (b'%PDF-1.4\n%mock\n', {'total_paginas': 1, 'paginas_ocr': 1, 'ocr_habilitado': True})

        resp = self.client.get(f'/sirel/documentos/processo/{processo.id}/integral/pdf/')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/pdf')
        self.assertIn('attachment; filename=', resp['Content-Disposition'])

    @patch('workflow.views._gerar_processo_etcm_zip_ou_pdf')
    def test_rota_geracao_etcm(self, mock_etcm):
        processo = Processo.objects.create(
            numero_processo_sirel='0051/2026',
            numero_processo_adm='ADM-0051',
            ano_referencia=2026,
            objeto='Processo para geração e-TCM',
        )
        mock_etcm.return_value = (b'%PDF-1.4\n%mock\n', 'arquivo.pdf', 'application/pdf', {'partes': 1})

        resp = self.client.get(f'/sirel/documentos/processo/{processo.id}/etcm/')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/pdf')
        self.assertIn('attachment; filename="arquivo.pdf"', resp['Content-Disposition'])

    def test_normalizacao_pdf_preserva_texto_e_paginacao(self):
        from io import BytesIO
        from pypdf import PdfReader

        base_pdf = _build_texto_pdf('Teste OCR', 'Linha 1 do documento\nLinha 2 do documento')
        normalizado, meta = _normalizar_pdf_para_padrao_documental(base_pdf, aplicar_ocr=True, pagina_max_kb=500)

        reader = PdfReader(BytesIO(normalizado))
        texto = reader.pages[0].extract_text() or ''

        self.assertEqual(len(reader.pages), 1)
        self.assertIn('Linha 1 do documento', texto)
        self.assertIn('Página 1 de 1', texto)
        self.assertEqual(meta.get('paginas_ocr', 0), 0)
        self.assertEqual(meta.get('paginas_texto_nativo', 0), 1)
