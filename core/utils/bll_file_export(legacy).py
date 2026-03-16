# -*- coding: utf-8 -*-
"""
Exportadores via ARQUIVOS (sem webservice) compatíveis com BLL:
- XLSX com abas: Lotes, Itens, TipoLance (igual ao Modelo_Global)
- CSV com seções: <PARTICIPANTES>, <PROCESSO>, <LOTES>, <VALORES UNITARIOS>, <CLASSIFICACAO>
"""
from __future__ import annotations
from decimal import Decimal
from typing import IO
from openpyxl import Workbook
from django.utils.timezone import localdate

from core.models import Processo, FornecimentoItem, Lote, ItemResultado


def export_bll_xlsx(processo: Processo, file_obj: IO[bytes]) -> None:
    wb = Workbook()
    ws_lotes = wb.active
    ws_lotes.title = "Lotes"
    ws_itens = wb.create_sheet("Itens")
    ws_tipo = wb.create_sheet("TipoLance")

    # Lotes
    ws_lotes.append(["Lote", "Título", "Tipo Lance", "Quantidade", "Margem Lance", "Garantia", "Local Entrega", "Exclusivo ME"])
    for lote in processo.lotes.all().order_by("numero"):
        ws_lotes.append([
            lote.numero, lote.titulo or f"Lote {lote.numero}", 2, 1, 0.01, "CONFORME EDITAL", "CONFORME EDITAL", "Não"
        ])

    # Itens
    ws_itens.append(["Lote", "Item", "Descrição", "Unidade", "Quantidade", "Valor Referência", "Info Detalhada", "Arquivo requerido"])
    for it in FornecimentoItem.objects.filter(processo=processo).select_related("lote").order_by("lote__numero","numero_item"):
        ws_itens.append([
            it.lote.numero if it.lote_id else 1,
            it.numero_item,
            it.descricao or "",
            it.unidade or "",
            float(it.quantidade or 0),
            float(it.valor_unitario or 0),
            "Não",
            "Não",
        ])

    # TipoLance
    ws_tipo.append(["idTipoLance", "Descrição"])
    ws_tipo.append([1, "Unitário"])
    ws_tipo.append([2, "Global"])
    ws_tipo.append([3, "Kit"])

    wb.save(file_obj)


def export_bll_csv(processo: Processo) -> str:
    def br(v):
        s = f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return s

    lines = []
    lines.append("<PARTICIPANTES>")
    # sem participantes no cadastro local (deixar vazio ou listar fornecedores de itens?)
    # aqui deixaremos vazio para BLL completar.

    lines.append("<PROCESSO>")
    lines.append(f"NÚMERO;{processo.numero_edital}-{processo.ano_referencia}")
    lines.append(f"DATA;{localdate().isoformat()}")

    lines.append("<LOTES>")
    for lote in processo.lotes.all().order_by("numero"):
        lines.append(f"{lote.numero};{lote.titulo or f'Lote {lote.numero}'};GLOBAL")

    lines.append("<VALORES UNITARIOS>")
    for it in FornecimentoItem.objects.filter(processo=processo).select_related("lote").order_by("lote__numero","numero_item"):
        lines.append(f"{it.lote.numero if it.lote_id else 1};{it.numero_item};{(it.descricao or '').replace(';',',')};{it.unidade or ''};{float(it.quantidade or 0)};{br(it.valor_unitario or 0)}")

    lines.append("<CLASSIFICACAO>")
    for r in ItemResultado.objects.filter(lote__processo=processo).select_related("lote","fornecedor").order_by("lote__numero","posicao"):
        doc = r.fornecedor.cnpj if r.fornecedor_id else ""
        lines.append(f"{r.lote.numero};{r.posicao};{doc};{br(r.valor_total or 0)}")

    return "\n".join(lines) + "\n"
