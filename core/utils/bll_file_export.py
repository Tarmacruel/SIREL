# -*- coding: utf-8 -*-
from __future__ import annotations

import unicodedata
from collections import defaultdict
from decimal import Decimal
from io import BytesIO

from django.utils.encoding import smart_str

from core.models import Processo, ProcessoItem


TIPO_LANCE_TO_ID = {
    "UNITARIO": 1,
    "UNITÁRIO": 1,
    "GLOBAL": 2,
    "KIT": 3,
}


def _normalize_text(value) -> str:
    txt = smart_str(value or "").replace("\r", " ").replace("\n", " ").strip()
    return txt


def _only_digits(value) -> str:
    return "".join(ch for ch in str(value or "") if ch.isdigit())


def _to_decimal(value) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def _tipo_lance_id(tipo_lance_raw: str, default: int = 2) -> int:
    tipo = _normalize_text(tipo_lance_raw).upper()
    return TIPO_LANCE_TO_ID.get(tipo, default)


def _texto_exclusivo_me(descricao: str) -> str:
    txt = unicodedata.normalize("NFKD", _normalize_text(descricao)).upper()
    txt = "".join(ch for ch in txt if not unicodedata.combining(ch))
    if "ME/EPP" in txt and ("EXCLUSIVO" in txt or "RESERVADA" in txt or "COTA" in txt):
        return "Sim"
    return "Não"


def _inferir_escopo(processo: Processo) -> str:
    escopo = _normalize_text(processo.escopo_disputa).upper()
    if escopo in {"GLOBAL", "LOTE", "ITEM"}:
        return escopo
    criterio = _normalize_text(processo.criterio_julgamento).upper()
    if "GLOBAL" in criterio:
        return "GLOBAL"
    if "LOTE" in criterio:
        return "LOTE"
    return "ITEM"


def _collect_items_base(processo: Processo):
    legacy_items = list(
        processo.itens.select_related("lote").all().order_by("lote__numero", "numero_item", "id")
    )
    if legacy_items:
        out = []
        for idx, it in enumerate(legacy_items, start=1):
            numero_externo = _only_digits(it.codigo_item_externo) or str(int(it.numero_item or idx))
            valor_ref = (
                _to_decimal(it.valor_unitario_estimado)
                if _to_decimal(it.valor_unitario_estimado) > 0
                else _to_decimal(it.valor_unitario)
            )
            out.append(
                {
                    "ordem": idx,
                    "lote_num_origem": int(it.lote.numero) if it.lote_id else None,
                    "lote_obj": it.lote,
                    "numero_item_origem": int(it.numero_item or idx),
                    "numero_item_externo": int(numero_externo),
                    "descricao": _normalize_text(it.descricao),
                    "unidade": _normalize_text(it.unidade),
                    "quantidade": _to_decimal(it.quantidade),
                    "valor_referencia": valor_ref,
                    "info_detalhada": "NÃO",
                    "arquivo_requerido": "NÃO",
                }
            )
        return out

    canonicos = list(ProcessoItem.objects.filter(processo=processo).order_by("numero_item"))
    out = []
    for idx, it in enumerate(canonicos, start=1):
        out.append(
            {
                "ordem": idx,
                "lote_num_origem": None,
                "lote_obj": None,
                "numero_item_origem": int(it.numero_item or idx),
                "numero_item_externo": int(it.numero_item or idx),
                "descricao": _normalize_text(it.descricao_snapshot),
                "unidade": _normalize_text(it.unidade_snapshot),
                "quantidade": _to_decimal(it.quantidade),
                "valor_referencia": _to_decimal(it.valor_referencia_unitario),
                "info_detalhada": "NÃO",
                "arquivo_requerido": "NÃO",
            }
        )
    return out


def _titulo_lote_default(numero_lote: int, descricao_item: str = "") -> str:
    desc = _normalize_text(descricao_item)
    if desc:
        return desc[:220]
    return f"LOTE {numero_lote}"


def _build_layout_rows(processo: Processo):
    escopo = _inferir_escopo(processo)
    base_items = _collect_items_base(processo)

    if not base_items:
        return {
            "lotes": [],
            "itens": [],
        }

    lotes_rows = []
    itens_rows = []

    if escopo == "GLOBAL":
        lote_num = 1
        titulo = _titulo_lote_default(lote_num, _normalize_text(processo.objeto))
        tipo_lance = 2
        exclusivo_me = "Sim" if all(_texto_exclusivo_me(i["descricao"]) == "Sim" for i in base_items) else "Não"
        lotes_rows.append(
            {
                "lote": lote_num,
                "titulo": titulo,
                "tipo_lance": tipo_lance,
                "quantidade": len(base_items),
                "margem_lance": Decimal("0.01"),
                "garantia": "CONFORME EDITAL",
                "local_entrega": "CONFORME EDITAL",
                "exclusivo_me": exclusivo_me,
            }
        )
        for idx, item in enumerate(base_items, start=1):
            itens_rows.append(
                {
                    "lote": lote_num,
                    "item": idx,
                    **item,
                }
            )
        return {"lotes": lotes_rows, "itens": itens_rows}

    if escopo == "ITEM":
        for lote_num, item in enumerate(base_items, start=1):
            lotes_rows.append(
                {
                    "lote": lote_num,
                    "titulo": _titulo_lote_default(lote_num, item["descricao"]),
                    "tipo_lance": 2,
                    "quantidade": 1,
                    "margem_lance": Decimal("0.01"),
                    "garantia": "CONFORME EDITAL",
                    "local_entrega": "CONFORME EDITAL",
                    "exclusivo_me": _texto_exclusivo_me(item["descricao"]),
                }
            )
            itens_rows.append(
                {
                    "lote": lote_num,
                    "item": 1,
                    **item,
                }
            )
        return {"lotes": lotes_rows, "itens": itens_rows}

    groups = defaultdict(list)
    for item in base_items:
        key = item["lote_num_origem"]
        if key is None:
            key = f"SEM_LOTE_{item['ordem']}"
        groups[key].append(item)

    group_keys = []
    numeric_keys = sorted([k for k in groups.keys() if isinstance(k, int)])
    text_keys = sorted([k for k in groups.keys() if not isinstance(k, int)], key=str)
    group_keys.extend(numeric_keys)
    group_keys.extend(text_keys)

    next_lote_num = (max(numeric_keys) + 1) if numeric_keys else 1
    for group_key in group_keys:
        items_group = groups[group_key]
        if isinstance(group_key, int):
            lote_num = group_key
        else:
            lote_num = next_lote_num
            next_lote_num += 1

        lote_obj = items_group[0]["lote_obj"]
        tipo_lance = _tipo_lance_id(getattr(lote_obj, "tipo_lance", ""), default=2)
        titulo = _normalize_text(getattr(lote_obj, "titulo", "")) if lote_obj else ""
        if not titulo:
            titulo = _titulo_lote_default(lote_num, items_group[0]["descricao"])
        exclusivo_me = "Sim" if all(_texto_exclusivo_me(i["descricao"]) == "Sim" for i in items_group) else "Não"

        lotes_rows.append(
            {
                "lote": lote_num,
                "titulo": titulo,
                "tipo_lance": tipo_lance,
                "quantidade": len(items_group),
                "margem_lance": Decimal("0.01"),
                "garantia": "CONFORME EDITAL",
                "local_entrega": "CONFORME EDITAL",
                "exclusivo_me": exclusivo_me,
            }
        )
        for idx, item in enumerate(sorted(items_group, key=lambda x: x["numero_item_externo"]), start=1):
            itens_rows.append(
                {
                    "lote": lote_num,
                    "item": idx,
                    **item,
                }
            )

    lotes_rows.sort(key=lambda r: r["lote"])
    itens_rows.sort(key=lambda r: (r["lote"], r["item"]))
    return {"lotes": lotes_rows, "itens": itens_rows}


def export_bll_csv(processo: Processo) -> str:
    raise RuntimeError("Exportação BLL em CSV desabilitada. Utilize apenas exportação XLSX.")


def export_bll_xlsx(processo: Processo, buffer: BytesIO):
    from openpyxl import Workbook

    layout = _build_layout_rows(processo)
    lotes_rows = layout["lotes"]
    itens_rows = layout["itens"]

    wb = Workbook()
    ws_lotes = wb.active
    ws_lotes.title = "LOTES"
    ws_lotes.append(
        ["LOTE", "TÍTULO", "TIPO LANCE", "QUANTIDADE", "MARGEM LANCE", "GARANTIA", "LOCAL ENTREGA", "EXCLUSIVO ME"]
    )
    for row in lotes_rows:
        ws_lotes.append(
            [
                int(row["lote"]),
                row["titulo"],
                int(row["tipo_lance"]),
                int(row["quantidade"]),
                float(row["margem_lance"]),
                row["garantia"],
                row["local_entrega"],
                row["exclusivo_me"],
            ]
        )

    ws_itens = wb.create_sheet("ITENS")
    ws_itens.append(
        ["LOTE", "ITEM", "DESCRIÇÃO", "UNID", "QUANTIDADE", "VALOR DE REFERÊNCIA", "INFO DETALHADA", "ARQUIVO REQUERIDO"]
    )
    for row in itens_rows:
        ws_itens.append(
            [
                int(row["lote"]),
                int(row["item"]),
                row["descricao"],
                row["unidade"],
                float(row["quantidade"]),
                float(row["valor_referencia"]),
                row["info_detalhada"] or "NÃO",
                row["arquivo_requerido"] or "NÃO",
            ]
        )

    ws_tipo = wb.create_sheet("TIPOLANCE")
    ws_tipo.append(["idTipoLance", "Descrição"])
    ws_tipo.append([1, "Unitário"])
    ws_tipo.append([2, "Global"])
    ws_tipo.append([3, "Kit"])

    wb.save(buffer)
