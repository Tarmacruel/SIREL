# -*- coding: utf-8 -*-
"""
Importação de arquivos BLL (CSV e XLSX) para Processo → Lote → Item.

Corrige:
- Repasse do status do LOTE para o ITEM quando a coluna de status do item não existir
- Vínculo do fornecedor por CNPJ (ou razão social se não houver CNPJ)
- Leitura e gravação de valor_unitario, proposta_inicial (se existir no modelo) e proposta_final
- Correção de acentos "estourados" (mojibake) comuns em CSV exportado

APIs públicas (usadas no admin):
- import_bll_csv_bytes(processo, raw_bytes)
- import_bll_xlsx_file(processo, filelike)
"""
from __future__ import annotations

import csv
import io
import re
import unicodedata
from decimal import Decimal
from typing import Dict, Any, List

from django.db import transaction

from core.utils.formatters import parse_brl
from core.models import Processo, Lote, FornecimentoItem, Fornecedor

# ---------------------- helpers ----------------------

def _strip_accents(s: str) -> str:
    return ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))

def _norm_header(s: str) -> str:
    s = (s or '').strip()
    s = _strip_accents(s).lower()
    s = re.sub(r'[^a-z0-9]+', '_', s)
    s = re.sub(r'_+', '_', s).strip('_')
    return s

def _digits(s: str) -> str:
    return re.sub(r'\D+', '', s or '')

def _fix_mojibake(text: str) -> str:
    """Tenta corrigir textos com acento "estourado" (latin-1 lido como utf-8 e vice-versa)."""
    if not text:
        return text
    try:
        x = text.encode('latin-1', errors='ignore').decode('utf-8', errors='ignore')
        if _seems_better(text, x):
            return x
    except Exception:
        pass
    return text

def _seems_better(orig: str, fixed: str) -> bool:
    if not fixed or fixed == orig:
        return False
    has_accent = any(ch in fixed for ch in "ÁÀÂÃÉÊÍÓÔÕÚÇáàâãéêíóôõúç")
    return has_accent or (abs(len(fixed) - len(orig)) <= 2)

def _to_decimal(val: Any) -> Decimal:
    if val is None:
        return Decimal('0')
    if isinstance(val, (int, float, Decimal)):
        return Decimal(str(val))
    s = str(val).strip()
    if not s:
        return Decimal('0')
    try:
        return parse_brl(s)
    except Exception:
        s2 = s.replace('.', '').replace(',', '.')
        try:
            return Decimal(s2)
        except Exception:
            return Decimal('0')

def _to_int(val: Any, default: int = 0) -> int:
    try:
        return int(str(val).strip())
    except Exception:
        try:
            return int(float(str(val).strip().replace(',', '.')))
        except Exception:
            return default

def _best(row: Dict[str, Any], *cands: str) -> Any:
    for c in cands:
        if c in row and row[c] not in (None, '', 'NULL', 'null', 'NaN'):
            return row[c]
    return None

# mapeamento de nomes possíveis → chave canônica
CANON = {
    'lote': ['lote', 'num_lote', 'numero_lote', 'n_lote'],
    'titulo_lote': ['titulo_lote', 'nome_lote', 'escopo', 'descricao_lote'],
    'status_lote': ['status_lote', 'situacao_lote', 'status_do_lote'],
    'numero_item': ['numero_item', 'item', 'n_item', 'num_item', 'sequencial_item'],
    'descricao': ['descricao', 'descricao_item', 'objeto', 'item_desc'],
    'unidade': ['unidade', 'unidade_medida', 'unid'],
    'quantidade': ['quantidade', 'qtd', 'qtde'],
    'valor_unitario': ['valor_unitario', 'vl_unitario', 'valor_un', 'vl_un', 'preco_unitario'],
    'proposta_inicial': ['proposta_inicial', 'lance_inicial', 'valor_proposta_inicial', 'primeiro_lance'],
    'proposta_final': ['proposta_final', 'lance_final', 'valor_proposta_final', 'melhor_lance', 'valor_homologado', 'valor_vencedor'],
    'fornecedor': ['fornecedor', 'razao_social', 'empresa', 'vencedor'],
    'cnpj': ['cnpj', 'cpf_cnpj', 'cnpj_cpf'],
    'status_item': ['status_item', 'situacao_item', 'status', 'resultado_item'],
}

def _normalize_row(raw_row: Dict[str, Any]) -> Dict[str, Any]:
    normed = { _norm_header(k): v for k, v in raw_row.items() }
    out = {}
    for kcanon, ks in CANON.items():
        for k in ks:
            if k in normed:
                out[kcanon] = normed[k]
                break
    out['_normed'] = normed
    return out

# ---------------------- CSV ----------------------

def _decode_csv(raw_bytes: bytes) -> str:
    for enc in ('utf-8-sig', 'utf-8', 'latin-1'):
        try:
            return raw_bytes.decode(enc)
        except Exception:
            continue
    return raw_bytes.decode('latin-1', errors='ignore')

def _rows_from_csv_text(text: str) -> List[Dict[str, Any]]:
    try:
        dialect = csv.Sniffer().sniff(text[:1024], delimiters=";,|\t")
        delim = dialect.delimiter
    except Exception:
        delim = ';' if text.count(';') >= text.count(',') else ','
    reader = csv.DictReader(io.StringIO(text), delimiter=delim)
    rows = []
    for r in reader:
        rows.append({k: (v.strip() if isinstance(v, str) else v) for k, v in r.items()})
    return rows

# ---------------------- XLSX ----------------------

def _rows_from_xlsx(filelike) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise RuntimeError("openpyxl não instalado. Rode: pip install openpyxl") from e
    wb = load_workbook(filelike, read_only=True, data_only=True)
    ws = wb.active
    headers = []
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c or '').strip() for c in row]
            continue
        rec = {}
        for h, c in zip(headers, row):
            rec[h] = '' if c is None else str(c).strip()
        rows.append(rec)
    return rows

# ---------------------- gravação no banco ----------------------

@transaction.atomic
def _import_rows(processo: Processo, rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    lotes_criados = 0
    itens_upd = 0
    fornecedores_criados = 0
    lotes_tocados = set()

    for raw in rows:
        row = _normalize_row(raw)

        # --- LOTE ---
        lote_num = _to_int(_best(row, 'lote'), default=1)
        titulo_lote = _fix_mojibake(_best(row, 'titulo_lote') or f"Lote {lote_num}")
        status_lote = _fix_mojibake(_best(row, 'status_lote') or '')

        lote, created_lote = Lote.objects.get_or_create(
            processo=processo,
            numero=lote_num,
            defaults={'titulo': titulo_lote, 'status': status_lote or ''}
        )
        if created_lote:
            lotes_criados += 1
        else:
            changed = False
            if titulo_lote and titulo_lote != (lote.titulo or ''):
                lote.titulo = titulo_lote; changed = True
            if status_lote and status_lote != (lote.status or ''):
                lote.status = status_lote; changed = True
            if changed:
                lote.save(update_fields=['titulo','status'])
        lotes_tocados.add(lote.pk)

        # --- FORNECEDOR ---
        fornecedor = None
        cnpj = _digits(_best(row, 'cnpj') or '')
        nome_forn = _fix_mojibake(_best(row, 'fornecedor') or '')
        if cnpj:
            fornecedor, _created = Fornecedor.objects.get_or_create(cnpj=cnpj, defaults={'razao_social': nome_forn[:255]})
            if _created: fornecedores_criados += 1
        elif nome_forn:
            fornecedor, _created = Fornecedor.objects.get_or_create(razao_social=nome_forn[:255], defaults={'cnpj': ''})
            if _created: fornecedores_criados += 1

        # --- ITEM ---
        num_item = _to_int(_best(row, 'numero_item'), default=0)
        descricao = _fix_mojibake(_best(row, 'descricao') or '')
        unidade = _best(row, 'unidade') or ''
        quantidade = _to_decimal(_best(row, 'quantidade') or '0')
        valor_unitario = _to_decimal(_best(row, 'valor_unitario') or '0')
        proposta_inicial = _to_decimal(_best(row, 'proposta_inicial') or '0')
        proposta_final = _to_decimal(_best(row, 'proposta_final') or '0')
        status_item = _fix_mojibake(_best(row, 'status_item') or '')

        # herda status do lote se não houver status do item
        if not status_item:
            status_item = status_lote

        obj, _created_item = FornecimentoItem.objects.update_or_create(
            processo=processo,
            lote=lote,
            numero_item=num_item,
            defaults=dict(
                descricao=descricao[:4000] if descricao else '',
                unidade=unidade[:50] if unidade else '',
                quantidade=quantidade,
                valor_unitario=valor_unitario,
                proposta_final=proposta_final,
                status_item=status_item[:100] if status_item else '',
                fornecedor=fornecedor
            )
        )
        if hasattr(obj, 'proposta_inicial'):
            if obj.proposta_inicial != proposta_inicial:
                obj.proposta_inicial = proposta_inicial
                obj.save(update_fields=['proposta_inicial'])
        itens_upd += 1

    # atualiza qtd_itens dos lotes tocados
    for lpk in lotes_tocados:
        Lote.objects.filter(pk=lpk).update(qtd_itens=FornecimentoItem.objects.filter(lote_id=lpk).count())

    return {
        "lotes_criados": lotes_criados,
        "itens_importados_ou_atualizados": itens_upd,
        "fornecedores_criados": fornecedores_criados,
    }

# ---------------------- APIs públicas ----------------------

def import_bll_csv_bytes(processo: Processo, raw_bytes: bytes) -> Dict[str, Any]:
    text = _decode_csv(raw_bytes)
    rows = _rows_from_csv_text(text)
    fixed_rows = []
    for r in rows:
        fixed_rows.append({k: _fix_mojibake(v) if isinstance(v, str) else v for k, v in r.items()})
    return _import_rows(processo, fixed_rows)

def import_bll_xlsx_file(processo: Processo, filelike) -> Dict[str, Any]:
    rows = _rows_from_xlsx(filelike)
    fixed_rows = []
    for r in rows:
        fixed_rows.append({k: _fix_mojibake(v) if isinstance(v, str) else v for k, v in r.items()})
    return _import_rows(processo, fixed_rows)
